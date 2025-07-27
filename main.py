import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os
import tempfile
import io

# Set page config
st.set_page_config(page_title="Vehicle Sales Analysis", page_icon="🚗", layout="wide")

# Title
st.title("🚗 Vehicle Sales Analysis Tool")
st.markdown("Upload your Excel files to generate comprehensive sales analysis reports")

# Logic of the tool
def drop_columns(data):
    drop_columns_list = ['Address','City','Locality','PinCode','Customer PhoneNo','Mobile No','Color Code','Color','Source',
                         'Manuf. Discount(-)','GatePass No.','GatePass Date','Registration Amount-RDTAX(+)',
                         'Insurance Amount-INSU(+)','Logistics Charges-HANDL(+)','Extended Warranty-EXTWAR(+)',
                         'Accessories Amount-ACCA','HSRP Charges-HSRP(+)','FASTAG Charges-FASTAG(+)','AMC Charges-AMC(+)',
                         'Other Charges-OTHCHG(+)','DISCOUNT ON INSURANC-DAT(-)','OP_SGST_RTO-OPSGSTRTO1(+)',
                         'OP_CGSTEV_RTO-OPCGSTEVRT1(+)','RTO CHARGES-RTO(+)','RDTAX (Paid)(-)','INSU (Paid)(-)','HANDL (Paid)(-)',
                         'EXTWAR (Paid)(-)','HSRP (Paid)(-)','FASTAG (Paid)(-)','AMC (Paid)(-)','OTHCHG (Paid)(-)',
                         'OP_SGST_RTO (Paid)(-)','OP_CGSTEV_RT (Paid)(-)','RTO (Paid)(-)','ACC_Paid(-)','Consumer Offer(Cash)',
                         'Consumer Offer(Acc)(-)','CorpDisc_Dealer','CorpDisc_Mfr(+)','Voucher Credit(+)','Voucher Debit(-)',
                         'InterestAmt(-)','Accessories Free Scheme Dlr Share','Accessories Free Scheme Mfr Share',
                         'Discounts on Insurance','EW Free Scheme Dlr Share','EW Free Scheme Mfr Share','Actual Acc. Amount Used(+)',
                         'Supplement Purchase Invoice No','Supplement Purchase Invoice Amount(-)','Debit Note No',
                         'Debit Note Amount(+)','Profit','Profit With Interest','DSA Adjustment']
    
    # Only drop columns that exist in the dataframe
    existing_columns = [col for col in drop_columns_list if col in data.columns]
    return data.drop(columns=existing_columns, axis=1)

def gst_calculation(data):
    data['gst'] = data['GST%'] + data['CESS%'] + 100
    return data

def additional_columns(data):
    additional_cols = [i for i in data.columns if "Additional" in i]
    for col in additional_cols:
        data[col + " "] = round(data[col] * 100 / data['gst'], 0)
    return data

def dlr_calculation(data):
    total = 0
    dlr_cols = [i for i in data.columns if "dlr" in i.lower() or "dealer" in i.lower()]
    for col in dlr_cols:
        data[col + " "] = round(data[col] * 100 / data['gst'], 0)
        total += data[col + " "]
    data['TOTAL DLR SHARE'] = total
    return data

def tata_share_calculation(data):
    total = 0
    tata_mfr_cols = [i for i in data.columns if "tata" in i.lower() or "mfr" in i.lower() or "mfg" in i.lower() or "manuf(+)" in i.lower()]
    for col in tata_mfr_cols:
        data[col + " "] = round(data[col] * 100 / data['gst'], 0)
        total += data[col + " "]
    data['TOTAL TATA SHARE'] = total
    return data

def fetching_discount_chassisno(data, sales_reco_data):
    dedup_sales = sales_reco_data.drop_duplicates(subset="Chassis_No")
    merged = pd.merge(data, 
                      dedup_sales[["Chassis_No", "Total Discount"]], 
                      left_on="ChassisNo", 
                      right_on="Chassis_No", 
                      how="inner")
    merged = merged.drop(columns=["Chassis_No"], axis=1)
    return merged

def purchase_sales(data):
    data['purchase -sales'] = round(data['Sale Price(+)'] - data['Purchase Price(-)'] - data['Total Discount'])
    return data

def margin_calculation(data):
    data['Margin'] = round(data['purchase -sales'] - data['AdditionalDiscount '] - data['TOTAL DLR SHARE'] - 
                           data['AdditionalFreeAcc(-) '] - data['DSAComission(-)'])
    return data

def total_row(data):
    total = data.loc[:len(data)-1].select_dtypes(include='number')
    total = total.drop(columns=['SNO'], axis=1, errors='ignore')
    total = total.sum()
    total_row = {}
    for col in data.columns:
        if col in total.index:
            total_row[col] = total[col]
        elif col.lower() == 'sno':
            total_row[col] = len(data)
        elif col.lower() == 'location':
            total_row[col] = f'Total ({len(data)})'
    total_row_df = pd.DataFrame([total_row], columns=list(total_row.keys()))
    data = pd.concat([data, total_row_df], ignore_index=True)
    data = data.rename(columns={'Total Discount': 'Tata DMS Credit'})
    return data

def chassis_file(data, file_path):
    data.to_excel(file_path, index=False, engine='openpyxl')
    wb = load_workbook(file_path)
    ws = wb.active
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    wb.save(file_path)

def chassis_file_trim(data, file_path):
    trim_data = data.loc[len(data)-1]
    new_data = data.copy()
    drop_column = []
    for col in data.columns:
        if trim_data[col] == 0:
            drop_column.append(col)
            new_data = new_data.drop(columns=col, axis=1)
    new_data.to_excel(file_path, index=False, engine='openpyxl')
    wb = load_workbook(file_path)
    ws = wb.active
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    wb.save(file_path)

def summary(data, file_path):
    sheet_name = 'Summary'
    row_spacing = 2
    location = data['Location'].unique()
    location = location[:len(location)-1]
    
    wb = load_workbook(file_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    wb.create_sheet(title=sheet_name)
    ws = wb[sheet_name]
    current_row = 1
    
    for i in location:
        show_rm = data[data['Location'] == i]
        model = show_rm['Model'].unique()
        
        sr_df1 = pd.DataFrame({
            'MODEL': model,
            'QTY':   [show_rm[show_rm['Model'] == m]['COUNT'].sum() for m in model],
            'SALE-PUR DIFF': [show_rm[show_rm['Model'] == m]['purchase -sales'].sum() for m in model],
            'Additional Discount': [show_rm[show_rm['Model'] == m]['AdditionalDiscount '].sum() for m in model],
            'Additional Accessories Discount': [show_rm[show_rm['Model'] == m]['AdditionalFreeAcc(-) '].sum() for m in model],
            'DSA Commission': [show_rm[show_rm['Model'] == m]['DSAComission(-)'].sum() for m in model],
            'Dlr share in Retail Support': [show_rm[show_rm['Model'] == m]['TOTAL DLR SHARE'].sum() for m in model],
            'Net Margin': [show_rm[show_rm['Model'] == m]['Margin'].sum() for m in model]
        })
        sr_df1['Per Car Margin'] = round(sr_df1['Net Margin'] / sr_df1['QTY'],0)
        total_row = round(sr_df1.select_dtypes(include='number').drop(columns='Per Car Margin').sum(),0)
        total_row['Per Car Margin'] = round(total_row['Net Margin'] / total_row['QTY'],0)
        total_row['MODEL'] = 'TOTAL'
        sr_df1.loc[len(sr_df1)] = total_row

        sr_df2 = pd.DataFrame({
            'TATA RETAIL SUPPORT': [show_rm[show_rm['Model'] == m]['TOTAL TATA SHARE'].sum() for m in model],
            'MFG share in Retail Support CREDIT IN TATA PUR': [show_rm[show_rm['Model'] == m]['Tata DMS Credit'].sum() for m in model]
        })
        total_row_2 = round(sr_df2.select_dtypes(include='number').sum(),0)
        sr_df2.loc[len(sr_df2)] = total_row_2

        sr_df3 = pd.DataFrame()
        sr_df3['TOTAL ADDL DISC'] = round(
            sr_df1['Additional Discount'] +
            sr_df1['Additional Accessories Discount'] +
            sr_df1['DSA Commission'],0
        )
        sr_df3['ADDIL DISC PER CAR'] = round(sr_df3['TOTAL ADDL DISC'] / sr_df1['QTY'],0)

        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            pd.DataFrame({f'Location: {i}': ['']}).to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=0, index=False, header=True)
            current_row += 2

            sr_df1.to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=0, index=False)
            sr_df2.to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=len(sr_df1.columns) + 2, index=False)
            sr_df3.to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=len(sr_df1.columns) + len(sr_df2.columns) + 4, index=False)

            current_row += max(len(sr_df1), len(sr_df2), len(sr_df3)) + row_spacing

    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2
    wb.save(file_path)

def verify_data(data, file_path):
    sheet_name = 'Difference'
    row_spacing = 2
    
    wb = load_workbook(file_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    wb.create_sheet(title=sheet_name)
    ws = wb[sheet_name]
    current_row = 1

    df1 = data[['Location','Sale Price(+)','Discount-DBT(-)','Purchase Price(-)']].groupby('Location').sum().round(0)
    purchase_price = df1['Purchase Price(-)']
    df1 = df1.drop(columns=['Purchase Price(-)'],axis=1)
    df1 = df1.rename(columns={'Sale Price(+)':'Sale','Discount-DBT(-)':'Discount'})
    df1['Net Sale'] = round(df1['Sale'] - df1['Discount'],0)
    df1['Purchase'] = purchase_price
    df1['Profit'] = round(df1['Net Sale'] - df1['Purchase'],0)
    
    df2 = data[['AdditionalDiscount ','TOTAL DLR SHARE','TOTAL TATA SHARE']].groupby(data['Location']).sum().round(0)
    df2['Total Discount'] = round(df2['AdditionalDiscount '] + df2['TOTAL DLR SHARE'] + df2['TOTAL TATA SHARE'],0)
    df2['Discount-DBT(-)'] = data['Discount-DBT(-)'].groupby(data['Location']).sum().round(0)
    df2['Difference'] = round(df2['Total Discount'] - df2['Discount-DBT(-)'],0)
    
    df3 = data[["TOTAL TATA SHARE","AdditionalFreeAcc(-) ",'DSAComission(-)','Tata DMS Credit']].groupby(data['Location']).sum().round(0)
    df3['Balance'] = round(df3['TOTAL TATA SHARE'] - df3['AdditionalFreeAcc(-) '] - df3['DSAComission(-)'] - df3['Tata DMS Credit'],0)
    total = round(df1['Profit'] + df3['Balance'],0)
    total = total.values
    total = total[len(total)-1]
    total_margin = data['Margin'].groupby(data['Location']).sum().round(0)
    total_margin = total_margin.values
    total_margin = total_margin[len(total_margin)-1]

    diff = abs(total-total_margin)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df1.to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=0, index=True)
        current_row += len(df1) + row_spacing

        df2.to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=0, index=True)
        current_row += len(df2) + row_spacing

        df3.to_excel(writer, sheet_name=sheet_name, startrow=current_row, startcol=0, index=True)
        current_row += len(df3) + row_spacing

        pd.DataFrame({f'Total': [total]}).to_excel(
            writer, sheet_name=sheet_name, startrow=current_row, startcol=len(df3.columns)-1, index=False, header=True
        )
        current_row += 2

        pd.DataFrame({f'Total Margin': [total_margin]}).to_excel(
            writer, sheet_name=sheet_name, startrow=current_row, startcol=len(df3.columns)-1, index=False, header=True
        )
        current_row += 2

        pd.DataFrame({f'Difference': [diff]}).to_excel(
            writer, sheet_name=sheet_name, startrow=current_row, startcol=len(df3.columns)-1, index=False, header=True
        )

    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2
    wb.save(file_path)

def get_sheet_names(file):
    """Get all sheet names from an Excel file"""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(file.read())
        tmp_path = tmp.name
    
    try:
        excel_file = pd.ExcelFile(tmp_path)
        sheet_names = excel_file.sheet_names
        excel_file.close()
        return sheet_names
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

def process_files(main_file, sales_reco_file, main_sheet_name, sales_sheet_name):
    # Create temporary files
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_main:
        tmp_main.write(main_file.read())
        main_file_path = tmp_main.name
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_sales:
        tmp_sales.write(sales_reco_file.read())
        sales_file_path = tmp_sales.name
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_output:
        output_file_path = tmp_output.name
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_trim:
        trim_file_path = tmp_trim.name
    
    try:
        # Process the data (your original logic)
        data = pd.read_excel(main_file_path, sheet_name=main_sheet_name, skiprows=6)
        sales_reco_data = pd.read_excel(sales_file_path, sheet_name=sales_sheet_name)
        data = drop_columns(data)
        data = gst_calculation(data)
        data['purchase -sales'] = 0
        data = additional_columns(data)
        data = dlr_calculation(data)
        data = tata_share_calculation(data)
        data['Margin'] = 0
        data = fetching_discount_chassisno(data, sales_reco_data)
        data = purchase_sales(data)
        data = margin_calculation(data)
        data = total_row(data)
        chassis_file(data, output_file_path)
        chassis_file_trim(data, trim_file_path)
        summary(data, output_file_path)
        verify_data(data, output_file_path)
        
        return output_file_path, trim_file_path
        
    except Exception as e:
        # Cleanup temp files in case of error
        for temp_file in [main_file_path, sales_file_path, output_file_path, trim_file_path]:
            if os.path.exists(temp_file):
                os.unlink(temp_file)
        raise e
    finally:
        # Cleanup input temp files
        for temp_file in [main_file_path, sales_file_path]:
            if os.path.exists(temp_file):
                os.unlink(temp_file)

# Streamlit UI
col1, col2 = st.columns(2)

with col1:
    st.subheader("📊 Main Data File")
    main_file = st.file_uploader(
        "Upload Book1.xlsx", 
        type=['xlsx'], 
        help="Upload the main sales data Excel file (Book1.xlsx)"
    )

with col2:
    st.subheader("💰 Margin Data File")
    sales_reco_file = st.file_uploader(
        "Upload MARGIN JULY 25.xlsx", 
        type=['xlsx'], 
        help="Upload the margin data Excel file (MARGIN JULY 25.xlsx)"
    )

# Sheet selection section
main_sheet_name = None
sales_sheet_name = None

if main_file:
    # Reset file pointer and get sheet names
    main_file.seek(0)
    main_sheets = get_sheet_names(main_file)
    
    if len(main_sheets) > 1:
        st.subheader("📋 Select Sheet from Main Data File")
        main_sheet_name = st.selectbox(
            "Choose the sheet to process from the main data file:",
            main_sheets,
            key="main_sheet"
        )
    else:
        main_sheet_name = main_sheets[0]
        st.info(f"📋 Main file has only one sheet: **{main_sheet_name}**")

if sales_reco_file:
    # Reset file pointer and get sheet names
    sales_reco_file.seek(0)
    sales_sheets = get_sheet_names(sales_reco_file)
    
    if len(sales_sheets) > 1:
        st.subheader("📋 Select Sheet from Margin Data File")
        sales_sheet_name = st.selectbox(
            "Choose the sheet to process from the margin data file:",
            sales_sheets,
            key="sales_sheet"
        )
    else:
        sales_sheet_name = sales_sheets[0]
        st.info(f"📋 Margin file has only one sheet: **{sales_sheet_name}**")

if main_file and sales_reco_file and main_sheet_name and sales_sheet_name:
    st.success("✅ Both files uploaded and sheets selected successfully!")
    
    if st.button("🚀 Process Files", type="primary", use_container_width=True):
        with st.spinner("Processing your files... This may take a few moments."):
            try:
                # Reset file pointers
                main_file.seek(0)
                sales_reco_file.seek(0)
                
                output_file_path, trim_file_path = process_files(main_file, sales_reco_file, main_sheet_name, sales_sheet_name)
                
                st.success("✨ Files processed successfully!")
                
                # Store file paths in session state for download
                st.session_state.output_file_path = output_file_path
                st.session_state.trim_file_path = trim_file_path
                st.session_state.files_processed = True
                        
            except Exception as e:
                st.error(f"❌ An error occurred while processing the files: {str(e)}")
                st.info("Please check that your files are in the correct format and try again.")

# Show download section if files are processed
if st.session_state.get('files_processed', False):
    st.markdown("---")
    st.subheader("📁 Download Your Files")
    st.markdown("**Customize your file names before downloading:**")
    
    col1, col2 = st.columns(2)
    
    with col1:
        complete_filename = st.text_input(
            "Complete Analysis File Name:",
            value="chassis",
            help="Enter the name for your complete analysis file (without .xlsx extension)"
        )
        
        if complete_filename:
            complete_filename = complete_filename.strip()
            if not complete_filename.endswith('.xlsx'):
                complete_filename += '.xlsx'
            
            try:
                with open(st.session_state.output_file_path, 'rb') as f:
                    st.download_button(
                        label=f"📥 Download {complete_filename}",
                        data=f.read(),
                        file_name=complete_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_complete"
                    )
            except:
                st.error("Error reading complete analysis file")
    
    with col2:
        trim_filename = st.text_input(
            "Trimmed Analysis File Name:",
            value="trim_chassis",
            help="Enter the name for your trimmed analysis file (without .xlsx extension)"
        )
        
        if trim_filename:
            trim_filename = trim_filename.strip()
            if not trim_filename.endswith('.xlsx'):
                trim_filename += '.xlsx'
            
            try:
                with open(st.session_state.trim_file_path, 'rb') as f:
                    st.download_button(
                        label=f"📥 Download {trim_filename}",
                        data=f.read(),
                        file_name=trim_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_trim"
                    )
            except:
                st.error("Error reading trimmed analysis file")
    
    # Cleanup button
    if st.button("🗑️ Clear and Process New Files", type="secondary", use_container_width=True):
        # Cleanup temp files
        for temp_file in [st.session_state.get('output_file_path'), st.session_state.get('trim_file_path')]:
            if temp_file and os.path.exists(temp_file):
                os.unlink(temp_file)
        
        # Clear session state
        for key in ['output_file_path', 'trim_file_path', 'files_processed']:
            if key in st.session_state:
                del st.session_state[key]
        
        st.rerun()

elif main_file and sales_reco_file:
    st.info("👆 Please select the appropriate sheets from both files to proceed.")
else:
    st.info("👆 Please upload both Excel files to begin processing.")

# Add information section
st.markdown("---")
st.subheader("📋 About This Tool")
st.markdown("""
This tool processes vehicle sales data and generates comprehensive analysis reports including:

- **Complete Analysis**: Full dataset with all calculations and metrics
- **Trimmed Analysis**: Filtered dataset removing columns with zero values
- **Summary Sheet**: Location-wise model analysis with margins
- **Difference Sheet**: Verification and reconciliation data

**File Requirements:**
- Main data file should be similar to `Book1.xlsx` format
- Margin file should contain a 'PV' sheet with chassis numbers and discount data
""")
